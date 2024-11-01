# coding=utf-8
from netmiko import ConnectHandler
from openpyxl import load_workbook
import os
from sys import exit
from netmiko import exceptions
import re


# 读取excel内设备列表信息
def check_and_get_dev_list(filename, sheet_name):
    excel_information = []
    sheet_header = []
    wb = load_workbook(filename)
    sh = wb[sheet_name]
    # 获取最大行数
    row = sh.max_row
    # 获取最大列数
    column = sh.max_column
    data = []
    # 获取表头写入列表中方便调用
    for data_1 in range(1, column + 1):
        get_sheet_header = sh.cell(row=1, column=data_1).value
        sheet_header.append(get_sheet_header)
    # 第一行为表头， 此处 row +1 是pyton循环时不读取最后一个数
    for row_1 in range(2, row + 1):
        # 存储一行信息
        sheet_data_1 = dict()
        # 逐行读取表中的数据
        for b in range(1, column + 1):
            cell = sh.cell(row=row_1, column=b).value
            # 将数据已字典形式写入 sheet_data_1 中
            # if cell != None:
            sheet_data_1[sheet_header[b - 1]] = cell
        excel_information.append(sheet_data_1)
    for i in excel_information:
        if i['ip'] != None:
            data.append(i)
    return data


# 获取excel数据并整合成dev字典
def get_dev():
    res = check_and_get_dev_list('./resource.xlsx', 'Sheet1')
    devices = []
    for i in res:
        if i['protocol'] == 'telnet':
            i['type'] = i['type'] + '_telnet'
        dev = {'device_type': i['type'],
               'host': i['ip'],
               'username': i['username'],
               'password': i['password'],
               'secret': i['enpassword'],
               'port': i['port'], }
        devices.append(dev)
    return devices


# 配置批量备份导出
def devices_confbak(devices=''):
    # 创建备份文件夹
    try:
        path = './conf_bak'
        os.makedirs(path)
    except FileExistsError:
        pass
    # 存储连接失败的IP
    failed_ips = []
    # 循环登录设备获取配置
    for dev in devices:
        try:
            with ConnectHandler(**dev) as conn:
                print('\n----------成功登录到：' + dev['host'] + '----------')
                conn.enable()
                if 'huawei' in dev['device_type']:
                    output = conn.send_command(command_string='dis current-configuration')
                elif 'ruijie_os' in dev['device_type']:
                    output = conn.send_command(command_string='show running-config')
                elif 'juniper' in dev['device_type']:
                    output1 = conn.send_command(command_string='show configuration')
                    output2 = conn.send_command(command_string='show configuration|display set')
                    output = 'show configuration:\n' + output1 + 'show configuration|display set:\n' + '\n' + output2
                else:
                    print('error')
                with open('./conf_bak/' + dev['host'] + '_conf_bak.txt', mode='w', encoding='utf8') as f:
                    print('正在备份:' + dev['host'])
                    # 文件读写异常处理
                    try:
                        f.write(output)
                    except PermissionError:
                        print('*****-无写入权限，请将文件夹赋予读写权限-*****')
                        continue
                    else:
                        print(dev['host']+'配置文件备份成功！')
        # 连接异常处理
        except exceptions.NetmikoAuthenticationException:
            print('\n**********' + dev['host'] + '：登录验证失败！**********')
            failed_ips.append(dev['host'])
            continue
        except exceptions.NetmikoTimeoutException:
            print('\n**********' + dev['host'] + '：目标不可达！**********')
            failed_ips.append(dev['host'])
            continue
        except exceptions.ReadTimeout:
            print('\n**********' + dev['host'] + '：读取超时，请检查enable密码是否正确！**********')
            failed_ips.append(dev['host'])
            continue
    if len(failed_ips) > 0:
        print('\n以下设备连接失败,请检查：')
        for x in failed_ips:
            print(x)
    return 1

# log文件导出
def devices_log(devices=''):
    # 创建备份文件夹
    try:
        path = './log'
        os.makedirs(path)
    except FileExistsError:
        pass
    # 存储连接失败的IP
    failed_ips = []
    # 循环登录设备获取配置
    for dev in devices:
        try:
            with ConnectHandler(**dev) as conn:
                print('\n----------成功登录到：' + dev['host'] + '----------')
                conn.enable()
                if 'huawei' in dev['device_type']:
                    output = conn.send_command(command_string='display logbuffer')
                elif 'ruijie_os' in dev['device_type']:
                    output = conn.send_command(command_string='show log')
                elif 'juniper' in dev['device_type']:
                    output = conn.send_command(command_string='show log messages')
                else:
                    print('error')
                with open('./log/' + dev['host'] + '_conf_log.txt', mode='w', encoding='utf8') as f:
                    print('正在备份:' + dev['host'])
                    # 文件读写异常处理
                    try:
                        f.write(output)
                    except PermissionError:
                        print('*****-无写入权限，请将文件夹赋予读写权限-*****')
                        continue
                    else:
                        print(dev['host']+'log文件备份成功！')
        # 连接异常处理
        except exceptions.NetmikoAuthenticationException:
            print('\n**********' + dev['host'] + '：登录验证失败！**********')
            failed_ips.append(dev['host'])
            continue
        except exceptions.NetmikoTimeoutException:
            print('\n**********' + dev['host'] + '：目标不可达！**********')
            failed_ips.append(dev['host'])
            continue
        except exceptions.ReadTimeout:
            print('\n**********' + dev['host'] + '：读取超时，请检查enable密码是否正确！**********')
            failed_ips.append(dev['host'])
            continue
    if len(failed_ips) > 0:
        print('\n以下设备连接失败,请检查：')
        for x in failed_ips:
            print(x)
    return 1

# 配置巡检
def devices_autocheck(devices='', cmd=''):
    # 存储命令执行回显
    results = []
    try:
        for x in range(len(devices)):
            # 循环登录设备
            with ConnectHandler(**devices[x]) as conn:
                conn.enable()
                print('正在巡检：' + devices[x]['host'] + ' ...')
                result = [devices[x]['host'], devices[x]['device_type']]
                for i in range(len(cmd)):
                    # 循环执行命令，根据不同设备执行不同命令
                    if 'huawei' in devices[x]['device_type']:
                        conn.send_command(command_string='sys', expect_string=']')
                        output = conn.send_command(command_string=str(cmd[i]['huawei']))
                    elif 'ruijie_os' in devices[x]['device_type']:
                        output = conn.send_command(command_string=str(cmd[i]['ruijie_os']))
                    elif 'juniper' in devices[x]['device_type']:
                        output = conn.send_command(command_string=str(cmd[i]['juniper']))
                    result.append(output)
                results.append(result)
    except exceptions.NetmikoAuthenticationException:
        print('\n**********' + devices[x]['host'] + '：登录验证失败！**********')
    except exceptions.NetmikoTimeoutException:
        print('\n**********' + devices[x]['host'] + '：目标不可达！**********')
    except exceptions.ReadTimeout:
        print('\n**********' + devices[x]['host'] + '：读取超时，请检查enable密码是否正确！**********')

    return results


# 运行主程序
if __name__ == '__main__':
    while True:
        print("\n##############################################\n")
        print("1：批量备份交换机配置")
        print("2：批量巡检交换机设备")
        print("0：退出")
        option = str(input("请输入需要的操作编号："))
        if option == '1':
            dev = get_dev()
            devices_confbak(devices=dev)
            devices_log(devices=dev)
            continue
        elif option == '2':
            # 定义巡检命令
            # cmds[x]['cisco']
            # cmds[x]['huawei']
            cmds_huawei = [
                {'cisco': 'show', 'huawei': 'display ip interface brief'},  # 查看设备三层接口状态及配置
                {'cisco': 'show', 'huawei': 'display interface description'},  # 查看设备接口描述信息
                {'cisco': 'show', 'huawei': 'display device'},  # 查看设备信息
                {'cisco': 'show', 'huawei': 'display alarm active'},  # 查看设备当前活动的告警信息
                {'cisco': 'show', 'huawei': 'display cpu-usage'},  # 查看设备各处理器使用情况
                {'cisco': 'show', 'huawei': 'display memory-usage'},  # 查看内存使用情况
                {'cisco': 'show', 'huawei': 'display ospf peer brief'},  # 查询ospf链路状态信息
                {'cisco': 'show', 'huawei': 'display ip routing-table'},  # 查看设备路由表信息
                {'cisco': 'show', 'huawei': 'display version'},  # 查看系统版本信息，运行时间
                {'cisco': 'show', 'huawei': 'display lldp neighbor brief'},  # 查看网络邻居信息


            ]
            cmds_ruijie = [
                {'ruijie_os': 'show ip interface brief', 'huawei': 'display ip interface brief'},  # 查看设备三层接口状态及配置
                {'ruijie_os': 'show interface status', 'huawei': 'display interface description'},  # 查询设备端口状态信息
                {'ruijie_os': 'show manuinfo', 'huawei': 'display device'},  # 查询设备硬件信息
                {'ruijie_os': 'show alarm', 'huawei': 'display alarm active'},  # 查询告警信息
                {'ruijie_os': 'show cpu', 'huawei': 'display cpu-usage'},  # 查看设备各处理器使用情况
                {'ruijie_os': 'show memory', 'huawei': 'display memory-usage'},  # 查看内存使用情况
                {'ruijie_os': 'show version', 'huawei': 'display version'},  # 查看系统版本信息，运行时间
                {'ruijie_os': 'show lldp neighbor', 'huawei': 'display lldp neighbor brief'},  # 查看网络邻居信息

            ]
            cmds_juniper = [
                {'ruijie_os': 'show ip interface brief', 'juniper': 'show chassis hardware'},  # 查看设备硬件状态
                {'ruijie_os': 'show interface status', 'juniper': 'show interfaces terse'},  # 查看所有物理接口状态
                {'ruijie_os': 'show manuinfo', 'juniper': 'show system uptime'},  # 查看系统启动时间
                {'ruijie_os': 'show alarm', 'juniper': 'show chassis alarms'},  # 查看设备硬件告警信息
                {'ruijie_os': 'show cpu', 'juniper': 'show system alarms'},  # 查看系统告警信息
                {'ruijie_os': 'show memory', 'juniper': 'show system processes summary'},  # 查看CPU使用使用率
                {'ruijie_os': 'show version', 'juniper': 'show system memory'},  # 查询系统内存使用
                {'ruijie_os': 'show lldp neighbor', 'juniper': 'show lldp neighbors'},  # 查看网络邻居信息


            ]
            dev = get_dev()
            for each in dev:
                # print(each['device_type'])
                if each['device_type'] == 'huawei':
                    checkres = devices_autocheck(dev, cmds_huawei)
                    for res in checkres:
                        fenge1 = '\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n'
                        output1 = res[0] + '-巡检结果：'
                        output2 = '\n设备三层接口状态及配置：\n' + res[2]
                        fenge2 = '\n=================================================================\n'
                        output3 = '设备端口状态信息：\n' + res[3]
                        output4 = '查询设备硬件信息：\n' + res[4]
                        if 'Invalid' in res[5]:
                            output5 = '设备不支持show alarm命令'
                        elif '' == res[5]:
                            output5 = '设备当前活动无告警！\n'
                        else:
                            output5 = '设备当前活动的告警信息：\n' + res[5]
                        output6 = '设备各处理器使用情况：\n' + res[6]
                        output7 = '内存利用率：\n' + res[7]
                        output8 = '系统版本信息，运行时间：\n' + res[8]
                        if 'Invalid' in res[9]:
                            output9 = '设备不支持show lldp neighbor命令'
                        else:
                            output9 = '查看网络邻居信息：\n' + res[9]
                        output_all = fenge1 + output1 + fenge2 + output2 + fenge2 + output3 + fenge2 + output4 + fenge2 + output5 + fenge2 + output6 + fenge2 + output7 + fenge2 + output8 + fenge2 + output9
                        #     print('\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
                        try:
                            path = './huaweixunjian/'
                            os.makedirs(path)
                        except FileExistsError:
                            pass
                        with open('./huaweixunjian/' + res[0] + '.txt', mode='w', encoding='utf8') as f:
                            print('正在写入巡检结果:' + res[0])
                            # 文件读写异常处理
                            try:
                                f.write(output_all)
                            except PermissionError:
                                print('*****-无写入权限，请将文件夹赋予读写权限-*****')
                                continue
                            else:
                                print('写入成功！')
                        with open('./conf_bak/' + res[0] + '_conf_bak.txt', 'r') as source_file:
                            source_content = source_file.read()
                            source_content = '\n=================================================================\n' + '设备配置信息：\n' + source_content + '\n=================================================================\n'
                        with open('./huaweixunjian/' + res[0] + '.txt', mode='a', encoding='utf8') as target_file:
                            target_file.write(source_content)

                    continue
                if each['device_type'] == 'ruijie_os':
                    checkres = devices_autocheck(dev, cmds_ruijie)
                    # print(checkres)
                    for res in checkres:
                        fenge1 = '\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n'
                        output1 = res[0] + '-巡检结果：'
                        output2 = '\n设备三层接口状态及配置：\n' + res[2]
                        fenge2 = '\n=================================================================\n'
                        output3 = '设备端口状态信息：\n' + res[3]
                        output4 = '查询设备硬件信息：\n' + res[4]
                        if 'Invalid' in res[5]:
                            output5 = '设备不支持show alarm命令'
                        elif '' == res[5]:
                            output5 = '设备当前活动无告警！\n'
                        else:
                            output5 = '设备当前活动的告警信息：\n' + res[5]
                        output6 = '设备各处理器使用情况：\n' + res[6]
                        output7 = '内存利用率：\n' + res[7]
                        output8 = '系统版本信息，运行时间：\n' + res[8]
                        if 'Invalid' in res[9]:
                            output9 = '设备不支持show lldp neighbor命令'
                        else:
                            output9 = '查看网络邻居信息：\n' + res[9]
                        output_all = fenge1 + output1 + fenge2 + output2 + fenge2 + output3 + fenge2 + output4 + fenge2 + output5 + fenge2 + output6 + fenge2 + output7 + fenge2 + output8 + fenge2 + output9
                        #     print('\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
                        try:
                            path = './ruijiexunjian/'
                            os.makedirs(path)
                        except FileExistsError:
                            pass
                        with open('./ruijiexunjian/' + res[0] + '.txt', mode='w', encoding='utf8') as f:
                            print('正在写入巡检结果:' + res[0])
                            # 文件读写异常处理
                            try:
                                f.write(output_all)
                            except PermissionError:
                                print('*****-无写入权限，请将文件夹赋予读写权限-*****')
                                continue
                            else:
                                print('写入成功！')
                        with open('./conf_bak/' + res[0] + '_conf_bak.txt', 'r') as source_file:
                            source_content = source_file.read()
                            source_content = '\n=================================================================\n' + '设备配置信息：\n' + source_content + '\n=================================================================\n'
                        with open('./ruijiexunjian/' + res[0] + '.txt', mode='a', encoding='utf8') as target_file:
                            target_file.write(source_content)

                    continue
                if each['device_type'] == 'juniper':
                    checkres = devices_autocheck(dev, cmds_juniper)
                    for res in checkres:
                        fenge1 = '\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n'
                        output1 = res[0] + '-巡检结果：'
                        output2 = '\n查看设备硬件状态：\n' + res[2]
                        fenge2 = '\n=================================================================\n'
                        output3 = '查看所有物理接口状态：\n' + res[3]
                        output4 = '查看系统启动时间：\n' + res[4]
                        output5 = '设备硬件告警信息：\n' + res[5]
                        output6 = '系统告警信息：\n' + res[6]
                        output7 = 'CPU使用使用率：\n' + res[7]
                        if 'Invalid' in res[8]:
                            output8 = '设备不支持show system memory命令'
                        else:
                            output8 = '系统内存使用：\n' + res[8]
                        output9 = '查看网络邻居信息：\n' + res[9]
                        output_all = fenge1 + output1 + fenge2 + output2 + fenge2 + output3 + fenge2 + output4 + fenge2 + output5 + fenge2 + output6 + fenge2 + output7 + fenge2 + output8 + fenge2 + output9
                        #     print('\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
                        try:
                            path = './juniperxunjian/'
                            os.makedirs(path)
                        except FileExistsError:
                            pass
                        with open('./juniperxunjian/' + res[0] + '.txt', mode='w', encoding='utf8') as f:
                            print('正在写入巡检结果:' + res[0])
                            # 文件读写异常处理
                            try:
                                f.write(output_all)
                            except PermissionError:
                                print('*****-无写入权限，请将文件夹赋予读写权限-*****')
                                continue
                            else:
                                print('写入成功！')
                        with open('./conf_bak/' + res[0] + '_conf_bak.txt', 'r') as source_file:
                            source_content = source_file.read()
                            source_content = '\n=================================================================\n' + '设备配置信息：\n' + source_content + '\n=================================================================\n'
                        with open('./juniperxunjian/' + res[0] + '.txt', mode='a', encoding='utf8') as target_file:
                            target_file.write(source_content)

                    continue

        elif option == '0':
            break
        else:
            print("请输入正确的编号！")